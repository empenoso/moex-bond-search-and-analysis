# 💰 Скачивание и обработка данных о денежном потоке облигаций 💰
#
# Этот Python скрипт автоматически скачивает данные о купонах и выплатах номинала
# через API Московской биржи для списка облигаций из Excel-файла bonds.xlsx и 
# записывает результат обратно в этот же файл.
#
# Установка зависимостей перед использованием: pip install requests openpyxl
#
# Автор: Михаил Шардин https://shardin.name/
# Дата создания: 29.01.2025
# Версия: 1.1
#
# Актуальная версия скрипта всегда здесь: https://github.com/empenoso/moex-bond-search-and-analysis
# 

import dataclasses
import logging
import os
import sys
from datetime import datetime

import requests
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

# Настройка кодировки для корректного вывода русского текста
if os.name == "nt":
    sys.stdout.reconfigure(encoding="utf-8")

log = logging.getLogger(__name__)
empty_log = logging.getLogger("empty")

def setup_logging():
    log.setLevel(logging.INFO)
    empty_log.setLevel(logging.INFO)

    handler = logging.StreamHandler(sys.stdout)
    empty_handler = logging.StreamHandler(sys.stdout)

    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    empty_formater = logging.Formatter("")

    handler.setFormatter(formatter)
    empty_handler.setFormatter(empty_formater)

    log.addHandler(handler)
    empty_log.addHandler(empty_handler)

@dataclasses.dataclass
class ExcelSheets:
    file_path: str
    workbook: Workbook
    data: Worksheet
    result: Worksheet


def main():
    setup_logging()
    excel_sheets = load_excel_file()
    excel_sheets = clean_excel_sheets_result(excel_sheets=excel_sheets)
    bonds = read_bonds(excel_sheets=excel_sheets)
    log.info(f"Считано {len(bonds)} облигаций для обработки.")
    cash_flow = process_bonds(bonds=bonds)
    write_data_to_excel(excel_sheets=excel_sheets, cache_flow=cash_flow)


def load_excel_file(file_path: str = "bonds.txt") -> ExcelSheets:
    # Загружаем Excel-файл
    file_path = "bonds.xlsx"
    wb = openpyxl.load_workbook(file_path)
    return ExcelSheets(file_path=file_path, workbook=wb, data=wb["Исходные данные"], result=wb["Ден.поток"])


def clean_excel_sheets_result(excel_sheets: ExcelSheets):
    # Очищаем лист с результатами
    column_names = ["Название", "Идентификатор", "Дата выплаты", "Денежный поток, ₽ (купон | выплата номинала)"]
    excel_sheets.result.delete_rows(1, excel_sheets.result.max_row)
    excel_sheets.result.append(column_names)
    return excel_sheets


def read_bonds(excel_sheets: ExcelSheets) -> list[tuple[str | float | datetime | None, ...]]:
    # Считываем данные из листа "Исходные данные"
    def is_not_empty_data(row) -> bool:
        return row[0] and row[1]
    
    data_iterator = excel_sheets.data.iter_rows(min_row=2, max_row=excel_sheets.data.max_row, values_only=True)
    return [row for row in data_iterator if is_not_empty_data(row)]


def process_bonds(bonds: list[tuple[str | float | datetime | None, ...]]) -> list[list[str]]:
    cash_flow = []
    # Обрабатываем каждую облигацию
    for ID, number in bonds:
        empty_log.info("")
        log.info(f"Обрабатываем {ID}, количество: {number} шт.")
        url = f"https://iss.moex.com/iss/statistics/engines/stock/markets/bonds/bondization/{ID}.json?iss.meta=off"
        log.info(f"Запрос к {url}")
        
        response = requests.get(url)
        json_data = response.json()
        
        assert isinstance(number, (float, int))
        cash_flow.extend(process_payment(json_data.get("amortizations", {}).get("data", []), number))
        coupons = json_data.get("coupons", {})
        cash_flow.extend(process_coupons(coupons.get("data", []), coupons.get("columns", []), number))

    return cash_flow


def process_coupons(coupons: list[tuple[str | int | float, ...]], columns: list[str], number: float | int) -> list[list[str]]:
    # Обработка купонов
    cash_flow = []

    isin_idx = columns.index("isin")
    name_idx = columns.index("name")
    coupondate_idx = columns.index("coupondate")
    value_rub_idx = columns.index("value_rub")

    for coupon in coupons:
        name = str(coupon[name_idx]).replace('"', '').replace("'", '').replace("\\", '')
        isin = coupon[isin_idx]
        coupon_date = coupon[coupondate_idx]

        # Преобразуем дату в объект datetime
        coupon_datetime = datetime.strptime(str(coupon_date), "%Y-%m-%d")

        if coupon_datetime > datetime.now():
            value_rub = float(coupon[value_rub_idx] or 0) * number
            flow = [f"{name} (купон 🏷️)", isin, coupon_datetime, value_rub]
            cash_flow.append(flow)
            log.info(f"Добавлен купон: {flow}")

    return cash_flow


def process_payment(amortizations: list[tuple[str | int | float, ...]], number: float | int) -> list[list[str]]:
    # Обработка выплат номинала
    cash_flow = []
    for amort in amortizations:
        name = str(amort[1]).replace('"', '').replace("'", '').replace("\\", '')
        isin = amort[0]
        amort_date = amort[3]

        # Преобразуем дату в объект datetime
        amort_datetime = datetime.strptime(str(amort_date), "%Y-%m-%d")

        if amort_datetime > datetime.now():
            value_rub = float(amort[9] or 0) * number
            flow = [f"{name} (номинал 💯)", isin, amort_datetime, value_rub]
            cash_flow.append(flow)
            log.info(f"Добавлена выплата номинала: {flow}")

    return cash_flow


def write_data_to_excel(excel_sheets: ExcelSheets, cache_flow: list[list[str]]):
    # Записываем данные в Excel
    for row in cache_flow:
        excel_sheets.result.append(row)

    # Устанавливаем формат ячеек
    for cell in excel_sheets.result["C"][1:]:  # Пропускаем заголовок
        cell.number_format = "DD.MM.YYYY"

    for cell in excel_sheets.result["D"][1:]:
        cell.number_format = '# ##0,00 ₽'

    # Добавляем запись об обновлении
    update_message = f"Данные автоматически обновлены {datetime.now().strftime('%d.%m.%Y в %H:%M:%S')}"
    excel_sheets.result.append(["", update_message])
    log.info(update_message)

    # Сохраняем изменения в файле
    excel_sheets.workbook.save(excel_sheets.file_path)
    log.info(f"Файл {excel_sheets.file_path} успешно обновлён.")
    log.info("Михаил Шардин https://shardin.name/\n")


if __name__ == "__main__":
    main()
    # В конце скрипта
    input("Нажмите клавишу Enter для выхода...")
