from datetime import datetime
from typing import cast

import openpyxl
import openpyxl.utils
from openpyxl.styles import Alignment
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from moex_bond_search_and_analysis.consts import DATETIME_FORMAT, MONTH_NAMES_RU_FULL
from moex_bond_search_and_analysis.logger import Logger
from moex_bond_search_and_analysis.schemas import Bond, ExcelSheets, SearchByCriteriaConditions


class ExcelSource:

    def __init__(self, filename: str) -> None:
        self.filename = filename

    def load_bonds(self) -> ExcelSheets:
        wb = openpyxl.load_workbook(self.filename)
        bonds = ExcelSheets(workbook=wb, data=wb["Исходные данные"], result=wb["Ден.поток"])
        # Очищаем лист с результатами
        column_names = ["Название", "Идентификатор", "Дата выплаты", "Денежный поток, ₽ (купон | выплата номинала)"]
        bonds.result.delete_rows(1, bonds.result.max_row)
        bonds.result.append(column_names)
        return bonds

    def write_bonds(self, sheets: ExcelSheets, cache_flow: list[list[str]], log: Logger):
        for row in cache_flow:
            sheets.result.append(row)

        # Устанавливаем формат ячеек
        for cell in sheets.result["C"][1:]:  # Пропускаем заголовок
            cell.number_format = "DD.MM.YYYY"  # TODO: использовать константу

        for cell in sheets.result["D"][1:]:
            cell.number_format = '# ##0,00 ₽'

        # Добавляем запись об обновлении
        update_message = f"Данные автоматически обновлены {datetime.now().strftime('%d.%m.%Y в %H:%M:%S')}"
        sheets.result.append(["", update_message])
        log.info(update_message)

        # Сохраняем изменения в файле
        sheets.workbook.save(self.filename)
        log.info(f"Файл {self.filename} успешно обновлён.")
    
    def write_search_by_criteria(self, data: list[Bond], conditions: SearchByCriteriaConditions, log: Logger) -> None:
        wb = openpyxl.Workbook()

        # Лист 'Результаты поиска'
        sheet_bonds = cast(Worksheet, wb.active)
        sheet_bonds.title = 'Результаты поиска'

        headers_bonds = ['Полное наименование', 'Код ценной бумаги', 'Нужна квалификация?', 'Цена, %', 'Объем сделок с 15 дней, шт.', 'Доходность', 'Дюрация, месяцев']
        headers_bonds.extend(MONTH_NAMES_RU_FULL)
        sheet_bonds.append(headers_bonds)

        # Форматирование столбца E как "# ##0"
        for cell in sheet_bonds["E"][1:]:
            cell.number_format = '# ##0'

        for bond in data:
            sheet_bonds.append(bond.as_list)

        # Центрирование данных на листе 'Результаты поиска'
        center_alignment = Alignment(horizontal='center')
        for row in sheet_bonds.iter_rows(min_row=1, max_row=sheet_bonds.max_row, min_col=1, max_col=sheet_bonds.max_column):  
            for cell in row:
                cell.alignment = center_alignment

        # Автоподбор ширины колонок для первого листа
        for column_cells in sheet_bonds.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            sheet_bonds.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].col_idx)].width = length + 2 

        # Фиксируем первую строку (заголовки)
        sheet_bonds.freeze_panes = 'A2'

        # Добавляем информацию об условиях поиска после таблицы
        last_row = sheet_bonds.max_row + 2  # Две строки после последней записи
        sheet_bonds.cell(
            row=last_row,
            column=1,
            value=f"Выборка сгенерирована {datetime.now().strftime(DATETIME_FORMAT)} по условиям:"
        )

        # Объединяем диапазон A:D для условий
        merge_range = f"A{last_row + 1}:D{last_row + 1}"
        sheet_bonds.merge_cells(merge_range)
        # Устанавливаем высоту строки (100)
        sheet_bonds.row_dimensions[last_row + 1].height = 100

        # Добавляем условия и настраиваем форматирование
        cell = sheet_bonds.cell(row=last_row + 1, column=1, value=conditions.as_string)
        cell.alignment = Alignment(wrap_text=True, vertical="top") # Перенос текста и выравнивание по верху
        self.__add_hiperlinks(sheet_bonds, last_row + 3)
        
        # Лист 'Лог'
        if log.messages:
            sheet_log = wb.create_sheet('Лог')
            sheet_log.title = 'Лог'
            sheet_log.column_dimensions['A'].width = 150  
            headers_log = ['Событие']
            sheet_log.append(headers_log)
            for log_entry in log.messages:  
                sheet_log.append([log_entry])

        wb.save(self.filename)
    
    def __add_hiperlinks(self, sheet: Worksheet, row_index: int) -> None:
        # Добавляем гиперссылки
        column = 1
        hyperlink_cell_author = cast(Cell, sheet.cell(row=row_index, column=column, value="Составил Михаил Шардин"))
        hyperlink_cell_author.hyperlink = "https://shardin.name/"
        hyperlink_cell_author.style = 'Hyperlink'

        hyperlink_cell_script = cast(Cell, sheet.cell(
            row=row_index + 1, column=column, value="Подробнее про скрипт поиска ликвидных облигаций в статье на GitHub"
        ))
        hyperlink_cell_script.hyperlink = "https://github.com/empenoso/moex-bond-search-and-analysis"
        hyperlink_cell_script.style = 'Hyperlink'
