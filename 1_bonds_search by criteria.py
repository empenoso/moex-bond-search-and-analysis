# 🕵️ Поиск ликвидных облигаций Мосбиржи по заданным критериям 🕵️
#
# Этот Python скрипт автоматически выполняет поиск облигаций, соответсвующих заданным
# критериям доходности, цены, дюрации и ликвидности, используя API Московской биржи.
# Результаты поиска, включающие информацию об облигациях и лог действий, 
# записываются в Excel-файл.
#
# Установка зависимостей перед использованием: pip install requests openpyxl asyncio
#
# Автор: Михаил Шардин https://shardin.name/
# Дата создания: 14.02.2025
# Версия: 1.3
#
# Актуальная версия скрипта всегда здесь: https://github.com/empenoso/moex-bond-search-and-analysis
#

import requests
import datetime
import openpyxl
import time
import inspect
import asyncio
import io
import sys  # Импорт sys для stdout
from openpyxl.styles import Alignment  # Добавляем Alignment для центрирования

# Глобальная переменная для задержки API запросов, 
# чтобы ожидать перед следующим запросом и соблюдать лимит в 50 запросов в минуту
API_DELAY = 1.2 # 1.2

class Tee(object):  # Класс Tee для дублирования вывода в консоль и в файл
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()  # Важно для немедленного отображения в консоли
    def flush(self):
        for f in self.files:
            f.flush()

def get_function_name():
    """Автоматически получаем имя функции."""
    return inspect.stack()[1][3]

async def start():
    """
    Основная функция запуска скрипта.
    Записываем время начала работы и вызывает функцию поиска облигаций.
    В конце выводит время выполнения скрипта.
    """
    start_time = datetime.datetime.now().timestamp() * 1000  # Unix Time Stamp - Epoch Converter
    print(f"🚀 Функция {get_function_name()} начала работу в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}. \n") 

    log_output = io.StringIO()  # Используем StringIO для захвата вывода
    original_stdout = sys.stdout  # Сохраняем стандартный вывод
    tee = Tee(original_stdout, log_output)  # Создаем Tee, чтобы писать и в консоль и в log_output
    sys.stdout = tee  # Перенаправляем стандартный вывод в tee

    moex_search_bonds_result, search_conditions = await moex_search_bonds() 

    sys.stdout = original_stdout  # Возвращаем стандартный вывод - важно вернуть именно original_stdout
    log_content = log_output.getvalue()  # Получаем захваченный лог
    excel_log_messages = log_content.splitlines()  # Разбиваем лог на строки для Excel

    curr_time = datetime.datetime.now().timestamp() * 1000
    duration = round((curr_time - start_time) / 1000 / 60 * 100) / 100  # время выполнения скрипта в минутах
    print(f"\n✅ Функция {get_function_name()} закончила работу в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}.") 
    print(f"⏳ Время выполнения {get_function_name()} в минутах: {duration}.") 

    # Записываем результаты в Excel файл
    excel_filename = f"bond_search_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    write_to_excel(moex_search_bonds_result, excel_log_messages, excel_filename, search_conditions)  # Передаем search_conditions 
    print(f"\n💾 Результаты записаны в Excel файл: {excel_filename}") 

    print("\nМихаил Шардин https://shardin.name/\n")
    
    # В конце скрипта
    input("Нажмите любую клавишу для выхода...")


async def moex_search_bonds():
    """
    Основная функция поиска облигаций по параметрам.
    Выполняет запросы к API Мосбиржи для поиска облигаций, соответствующих заданным критериям.
    Возвращает список найденных облигаций, список сообщений лога и условия поиска. 
    """
    yield_more = 15  # Доходность больше этой цифры
    yield_less = 40  # Доходность меньше этой цифры
    price_more = 70  # Цена больше этой цифры
    price_less = 120  # Цена меньше этой цифры
    duration_more = 3  # Дюрация больше этой цифры
    duration_less = 18  # Дюрация меньше этой цифры
    volume_more = 2000  # Объем сделок в каждый из n дней, шт. больше этой цифры
    bond_volume_more = 60000  # Совокупный объем сделок за n дней, шт. больше этой цифры
    offer_yes_no = "ДА"  # Учитывать, чтобы денежные выплаты были известны до самого погашения?
    # ДА - облигации только с известными цифрами выплаты купонов
    # НЕТ - не важно, пусть в какие-то даты вместо выплаты прочерк
    conditions = f"""{yield_more}% < Доходность < {yield_less}%
{price_more}% < Цена < {price_less}%
{duration_more} мес. < Дюрация < {duration_less} мес.
Значения всех купонов известны до самого погашения: {offer_yes_no}.
Объем сделок в каждый из 15 последних дней (c {(datetime.datetime.now() - datetime.timedelta(days=15)).strftime('%d.%m.%Y')}) > {volume_more} шт.
Совокупный объем сделок за 15 дней больше {bond_volume_more} шт.
Поиск в Т0, Т+, Т+ (USD) - Основной режим - безадрес.
""" # Условия для вывода в Excel
    bonds = []
    count = 0
    error_counter = 0  # Счётчик ошибок соединений с серверами Московской биржи

    board_groups = [58, 193, 105, 77, 207, 167, 245]  # https://iss.moex.com/iss/engines/stock/markets/bonds/boardgroups/
    for t in board_groups:
        url = f"https://iss.moex.com/iss/engines/stock/markets/bonds/boardgroups/{t}/securities.json?iss.dp=comma&iss.meta=off&iss.only=securities,marketdata&securities.columns=SECID,SECNAME,PREVLEGALCLOSEPRICE&marketdata.columns=SECID,YIELD,DURATION"
        print(f"🔗 {get_function_name()}. Ссылка поиска всех доступных облигаций группы: {url}.") 

        time.sleep(API_DELAY)  # Задержка API_DELAY с между запросами

        try:
            response = requests.get(url)
            response.raise_for_status()  # Проверка на HTTP ошибки
            json_data = response.json()
        except requests.exceptions.RequestException as e:
            error_counter += 1
            print(f"⚠️ Ошибка при запросе к API: {e}") 
            continue  # Переходим к следующей группе, если произошла ошибка

        if not json_data or not json_data.get('marketdata') or not json_data['marketdata'].get('data'):
            print(f'📉 {get_function_name()}. Нет данных c Московской биржи для группы {t}. Проверьте вручную по ссылке выше.') 
            continue  # Переходим к следующей группе, если нет данных

        bond_list = json_data['securities']['data']
        count = len(bond_list)
        print(f'📃 {get_function_name()}. Всего в списке группы {t}: {count} бумаг.\n') 

        market_data = json_data['marketdata']['data']
        market_data_dict = {item[0]: item for item in market_data if item}  # Создаем словарь для быстрого доступа к данным marketdata по SECID

        for i in range(count):
            # если из-за сетевой ошибки цикл прервался, тогда повтор
            retry_count = 0  # Счётчик попыток
            while retry_count < 5:  # Лимит перезапуска до 5 раз
                try:
                    bond_name = bond_list[i][1].replace('"', '').replace("'", '')
                    secid = bond_list[i][0]
                    bond_price = bond_list[i][2]

                    bond_market_data = market_data_dict.get(secid)
                    if not bond_market_data:
                        print(f"❌ {get_function_name()} в {datetime.datetime.now().strftime('%H:%M:%S')}. Строка {i + 1} из {count}: {bond_name} ({secid}): Данные о доходности и дюрации отсутствуют.") 
                        retry_count = 5
                        continue

                    bond_yield = bond_market_data[1]
                    bond_duration = bond_market_data[2] / 30 if bond_market_data[2] else 0  # кол-во оставшихся месяцев, делим на 30 если есть значение, иначе 0
                    bond_duration = round(bond_duration * 100) / 100

                    print(f"🔎 {get_function_name()} в {datetime.datetime.now().strftime('%H:%M:%S')}. Строка {i + 1} из {count}: {bond_name} ({secid}): цена={bond_price}%, доходность={bond_yield}%, дюрация={bond_duration} мес.") 

                    if (bond_yield is not None and yield_more <= bond_yield <= yield_less and  # условия выборки
                        bond_price is not None and price_more <= bond_price <= price_less and 
                        duration_more < bond_duration < duration_less):
                        print(f"✅ {get_function_name()}.   \\-> Условие доходности ({yield_more} < {bond_yield}% < {yield_less}), цены ({price_more} < {bond_price}% < {price_less}) и дюрации ({duration_more} < {bond_duration} мес. < {duration_less}) для {bond_name} прошло.") 
                        volume_data = await moex_search_volume(secid, volume_more)
                        bond_volume = volume_data['value']
                        print(f"📊 {get_function_name()}. \\-> Совокупный объем сделок за n дней: {bond_volume}, а условие {bond_volume_more} шт.") 
                        if volume_data['low_liquid'] == 0 and bond_volume > bond_volume_more:  # lowLiquid: 0 и 1 - переключатели.
                            # ❗ 0 - чтобы оборот был строго больше заданного
                            # ❗ 1 - фильтр оборота не учитывается, в выборку попадают все бумаги, подходящие по остальным параметрам
                            payments_data = await moex_search_months_of_payments(secid)
                            months_of_payments_null = payments_data['value_rub_null']
                            is_qualified_investors = await moex_search_is_qualified_investors(secid)
                            if offer_yes_no == "ДА" and months_of_payments_null == 0:
                                bonds.append([bond_name, secid, is_qualified_investors, bond_price, bond_volume, bond_yield, bond_duration, payments_data['months_payment_marks']])  
                                print(f"🗓️ {get_function_name()}. Для {bond_name} ({secid}) все даты будущих платежей с известным значением выплат.") 
                                print(f'⭐ {get_function_name()}. Результат № {len(bonds)}: {bonds[-1]}.') 
                            elif offer_yes_no == "НЕТ":
                                bonds.append([bond_name, secid, is_qualified_investors, bond_price, bond_volume, bond_yield, bond_duration, payments_data['months_payment_marks']])  
                                print(f'⭐ {get_function_name()}. Результат № {len(bonds)}: {bonds[-1]}.\n') 
                            else:
                                print(f"🚫 {get_function_name()}. Облигация {bond_name} ({secid}) в выборку не попадает из-за того, что есть даты когда значения выплат неизвестны.\n") 
                        else:
                            print(f"💧 {get_function_name()}. Облигация {bond_name} ({secid}) в выборку не попадает из-за малых оборотов или доступно мало торговых дней.\n") 
                    else:
                         print(f'⏭️ {get_function_name()} Пропуск {secid}: не соответствует базовым параметрам.\n') 
                    retry_count = 5  # Успешное завершение группы, прерываем повторение

                except requests.exceptions.RequestException as e:
                    retry_count += 1
                    error_counter += 1
                    print(f"\n⚠️ Ошибка при обработке строки {i + 1}: {e}.\n🔄 Попытка {retry_count} из 5. Ожидание 60 секунд.\n") 
                    time.sleep(60)  # Ожидание перед повтором
                except Exception as e:  # Обработка любых других ошибок
                    retry_count += 1
                    error_counter += 1
                    print(f"\n🔥 Непредвиденная ошибка при обработке строки {i + 1}: {e}.\n🔄 Попытка {retry_count} из 5. Ожидание 60 секунд.\n") 

    if not bonds:
        print(f"📭 {get_function_name()}. В массиве нет строк.") 
        return conditions 

    bonds.sort(key=lambda x: x[4], reverse=True)  # сортировка по столбцу Объем сделок за n дней, шт.

    if bonds:
        print(f"📊 {get_function_name()}. Начало выборки: {bonds[0]}, ...") 
    print(f"🐞 {get_function_name()}. Количество ошибок в соединении с Московской биржей: {error_counter}, все данные получены.") 

    return bonds, conditions


async def moex_search_volume(security_id, threshold_value):
    """
    Объем сделок в каждый из n дней больше определенного порога.
    Получает данные об объемах торгов для заданной облигации за последние 15 дней.
    Возвращает словарь с информацией о ликвидности, суммарном объеме и сообщениями лога.
    """
    now = datetime.datetime.now()
    date_request_previous = (now - datetime.timedelta(days=15)).strftime('%Y-%m-%d')  # этот день n дней назад
    board_id = await moex_board_id(security_id)
    if not board_id:
        print(f"⚠️ Не удалось получить board_id для {security_id}. Поиск объема прерван.") 
        return {'low_liquid': 1, 'value': 0, 'log': []}  # Лог теперь захватывается stdout, возвращаем пустой список

    url = f"https://iss.moex.com/iss/history/engines/stock/markets/bonds/boards/{board_id}/securities/{security_id}.json?iss.meta=off&iss.only=history&history.columns=SECID,TRADEDATE,VOLUME,NUMTRADES&limit=20&from={date_request_previous}"
    # numtrades - Минимальное количество сделок с бумагой
    # VOLUME - оборот в количестве бумаг (Объем сделок, шт)
    print(f'🔗 {get_function_name()}. Ссылка для поиска объёма сделок {security_id}: {url}') 
    try:
        time.sleep(API_DELAY)  # Задержка API_DELAY с между запросами

        response = requests.get(url)
        response.raise_for_status()
        json_data = response.json()
        history_data = json_data['history']['data']

        count = len(history_data)
        volume_sum = 0
        low_liquid = 0
        for i in range(count):
            volume = history_data[i][2]
            volume_sum += volume
            if threshold_value > volume:  # если оборот в конкретный день меньше
                low_liquid = 1
                print(f"📉 {get_function_name()}. На {i + 1}-й день ({history_data[i][1]}) из {count} оборот по бумаге {security_id} меньше чем {threshold_value}: {volume} шт.") 
            if count < 6:  # если всего дней в апи на этом периоде очень мало
                low_liquid = 1
                print(f"⚠️ {get_function_name()}. Всего в АПИ Мосбиржи доступно {count} дней, а надо хотя бы больше 6 торговых дней с {date_request_previous}!") 

        if low_liquid != 1:
            print(f"📈 {get_function_name()}. Во всех {count} днях оборот по бумаге {security_id} был больше, чем {threshold_value} шт каждый день.") 
        print(f"📊 {get_function_name()}. Итоговый оборот в бумагах (объем сделок, шт) за {count} дней: {volume_sum} шт нарастающим итогом.") 
        return {
            'low_liquid': low_liquid,
            'value': volume_sum
        }
    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка c {security_id} в {get_function_name()}: {e}") 
        return {'low_liquid': 1, 'value': 0} 
    except Exception as e:
        print(f"🔥 Непредвиденная ошибка c {security_id} в {get_function_name()}: {e}") 
        return {'low_liquid': 1, 'value': 0} 


async def moex_board_id(security_id):
    """
    Узнаем boardid любой бумаги по тикеру.
    Получает board_id для заданной облигации.
    Возвращает board_id или None в случае ошибки.
    """
    url = f"https://iss.moex.com/iss/securities/{security_id}.json?iss.meta=off&iss.only=boards&boards.columns=secid,boardid,is_primary"
    try:
        time.sleep(API_DELAY)  # Задержка API_DELAY с между запросами

        response = requests.get(url)
        response.raise_for_status()
        json_data = response.json()

        board_id_data = json_data['boards']['data']
        primary_board = next((board[1] for board in board_id_data if board[2] == 1), None)  # Находим board_id где is_primary = 1

        if primary_board:
            return primary_board
        else:
            print(f"⚠️ Не найден primary board_id для {security_id}.") 
            return None

    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка c {security_id} в {get_function_name()}: {e}") 
        return None
    except Exception as e:
        print(f"🔥 Непредвиденная ошибка c {security_id} в {get_function_name()}: {e}") 
        return None


async def moex_search_months_of_payments(security_id):
    """
    Узнаём месяцы, когда происходят выплаты.
    Получает данные о купонных выплатах для заданной облигации.
    Возвращает словарь с информацией о месяцах выплат, наличии неизвестных выплат и months_payment_marks.
    """
    url = f"https://iss.moex.com/iss/statistics/engines/stock/markets/bonds/bondization/{security_id}.json?iss.meta=off&iss.only=coupons&start=0&limit=100"
    print(f'🔗 {get_function_name()}. Ссылка для поиска месяцев выплат для {security_id}: {url}.') 
    try:
        time.sleep(API_DELAY)  # Задержка API_DELAY с между запросами

        response = requests.get(url)
        response.raise_for_status()
        json_data = response.json()

        coupon_data = json_data['coupons']['data']

        coupon_dates = []
        value_rub_null = 0
        for i in range(len(coupon_data)):
            coupondate = coupon_data[i][3]  # даты купона
            value_rub = coupon_data[i][9]  # сумма выплаты купона
            in_future = datetime.datetime.strptime(coupondate, '%Y-%m-%d') > datetime.datetime.now()
            if in_future:
                coupon_dates.append(int(coupondate.split("-")[1]))  # Добавляем номер месяца
                if value_rub is None:
                    value_rub_null += 1

        if value_rub_null > 0:
            print(f"⚠️ {get_function_name()}. Для {security_id} есть {value_rub_null} дат(ы) будущих платежей с неизвестным значением выплат.") 

        unique_dates = sorted(list(set(coupon_dates)))  # уникальные значения месяцев и сортировка
        print(f"🗓️ {get_function_name()}. Купоны для {security_id} выплачиваются в {unique_dates} месяцы.") 

        month_names_short_ru = ['янв', 'фев', 'мар', 'апр', 'май', 'июн', 'июл', 'авг', 'сен', 'окт', 'ноя', 'дек']
        months_payment_marks = {}  # Словарь для отметок месяцев
        for month_num in range(1, 13):
            months_payment_marks[month_names_short_ru[month_num-1]] = "✅" if month_num in unique_dates else ""  # Отмечаем месяцы с выплатами

        return {
            'value_rub_null': value_rub_null,
            'months_payment_marks': months_payment_marks  # Возвращаем словарь с отметками
        }

    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка c {security_id} в {get_function_name()}: {e}") 
        return  {'value_rub_null': 0, 'months_payment_marks': {}}  
    except Exception as e:
        print(f"🔥 Непредвиденная ошибка c {security_id} в {get_function_name()}: {e}") 
        return  {'value_rub_null': 0, 'months_payment_marks': {}}  


async def moex_search_is_qualified_investors(security_id):
    """
    Определяем это бумага для квалифицированных инвесторов или нет.
    Получает информацию о необходимости квалификации для покупки облигации.
    Возвращает 'да' или 'нет'.
    """
    url = f"https://iss.moex.com/iss/securities/{security_id}.json?iss.meta=off&iss.only=description&description.columns=name,title,value"
    print(f'🔗 {get_function_name()}. Ссылка для поиска общей информации по {security_id}: {url}') 
    try:

        time.sleep(API_DELAY)  # Задержка API_DELAY с между запросами

        response = requests.get(url)
        response.raise_for_status()
        json_data = response.json()
        description_data = json_data['description']['data']

        is_qualified_investors_data = next((item for item in description_data if item[0] == 'ISQUALIFIEDINVESTORS'), None)
        qual_investor_group_data = next((item for item in description_data if item[0] == 'QUALINVESTORGROUP'), None)

        is_qualified_investors = int(is_qualified_investors_data[2]) if is_qualified_investors_data and is_qualified_investors_data[2] else 0  # По умолчанию 0, если не найдено
        qual_investor_group = qual_investor_group_data[2] if qual_investor_group_data and qual_investor_group_data[2] else "не определена"  # Текст по умолчанию, если не найден

        if is_qualified_investors == 0:
            print(f"👤 {get_function_name()}. Для {security_id} квалификация для покупки НЕ нужна.") 
            return 'нет'
        else:
            print(f"👨‍💼 {get_function_name()}. {security_id} это бумага для квалифицированных инвесторов категории: \"{qual_investor_group}\"") 
            return 'да'

    except requests.exceptions.RequestException as e:
        print(f"⚠️ Ошибка c {security_id} в {get_function_name()}: {e}") 
        return 'ошибка'  # Return some error indicator
    except Exception as e:
        print(f"🔥 Непредвиденная ошибка c {security_id} в {get_function_name()}: {e}") 
        return 'ошибка'  # Return some error indicator


def write_to_excel(bonds, excel_log_messages, filename, conditions):
    """
    Записывает результаты поиска облигаций и лог в Excel файл.
    Создает две вкладки: 'Результаты поиска' и 'Лог'.
    """
    wb = openpyxl.Workbook()

    # Лист 'Результаты поиска'
    sheet_bonds = wb.active
    sheet_bonds.title = 'Результаты поиска'
    headers_bonds = ['Полное наименование', 'Код ценной бумаги', 'Нужна квалификация?', 'Цена, %', 'Объем сделок с 15 дней, шт.', 'Доходность', 'Дюрация, месяцев']
    month_names_full_ru = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']  # Полные названия месяцев для заголовков
    headers_bonds.extend(month_names_full_ru)  # Добавляем заголовки месяцев
    sheet_bonds.append(headers_bonds)

    sheet_bonds = wb["Результаты поиска"]

    # Форматирование столбца E как "# ##0"
    for cell in sheet_bonds["E"][1:]:
       cell.number_format = '# ##0'

    # volume_column = sheet_bonds['E']
    # number_format = "#,##0"
    # for cell in volume_column:
    #     cell.number_format = number_format        

    for bond_data in bonds:
        bond_row = bond_data[:7]  # Первые 7 элементов - основная информация
        payment_marks = bond_data[7]  # Словарь с отметками месяцев
        month_marks_list = [payment_marks.get(month, "") for month in month_names_short_ru]  # Получаем отметки в порядке месяцев
        bond_row.extend(month_marks_list)  # Добавляем отметки месяцев в строку
        sheet_bonds.append(bond_row)

    # Центрирование данных на листе 'Результаты поиска'
    center_alignment = Alignment(horizontal='center')
    for row in sheet_bonds.iter_rows(min_row=1, max_row=sheet_bonds.max_row, min_col=1, max_col=sheet_bonds.max_column):  
        for cell in row:
            cell.alignment = center_alignment

    # Автоподбор ширины колонок для первого листа
    for column_cells in sheet_bonds.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        sheet_bonds.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].col_idx)].width = length + 2 

    # Фиксируем первую строку (заголовки)
    sheet_bonds.freeze_panes = 'A2'

    # Добавляем информацию об условиях поиска после таблицы
    last_row = sheet_bonds.max_row + 2  # Две строки после последней записи
    sheet_bonds.cell(row=last_row, column=1, value=f"Выборка сгенерирована {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')} по условиям:")

    # Объединяем диапазон A:D для условий
    merge_range = f"A{last_row + 1}:D{last_row + 1}"
    sheet_bonds.merge_cells(merge_range)
    # Устанавливаем высоту строки (100)
    sheet_bonds.row_dimensions[last_row + 1].height = 100

    # Добавляем условия и настраиваем форматирование
    cell = sheet_bonds.cell(row=last_row + 1, column=1, value=conditions)
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top") # Перенос текста и выравнивание по верху

    # Добавляем гиперссылки
    hyperlink_row = last_row + 3  # Строка для гиперссылок
    hyperlink_cell_author = sheet_bonds.cell(row=hyperlink_row, column=1)
    hyperlink_cell_script = sheet_bonds.cell(row=hyperlink_row + 1, column=1)

    hyperlink_cell_author.value = "Составил Михаил Шардин" 
    hyperlink_cell_script.value = "Подробнее про скрипт поиска ликвидных облигаций в статье на GitHub" 

    hyperlink_cell_author.hyperlink = "https://shardin.name/"
    hyperlink_cell_author.style = 'Hyperlink'

    hyperlink_cell_script.hyperlink = "https://github.com/empenoso/moex-bond-search-and-analysis"
    hyperlink_cell_script.style = 'Hyperlink'

    # Лист 'Лог'
    sheet_log = wb.create_sheet('Лог')
    sheet_log.title = 'Лог'
    sheet_log.column_dimensions['A'].width = 150  
    headers_log = ['Событие']
    sheet_log.append(headers_log)
    for log_entry in excel_log_messages:  
        sheet_log.append([log_entry])

    wb.save(filename)

month_names_short_ru = ['янв', 'фев', 'мар', 'апр', 'май', 'июн', 'июл', 'авг', 'сен', 'окт', 'ноя', 'дек']  # Короткие названия месяцев для порядка столбцов

if __name__ == "__main__":
    asyncio.run(start())